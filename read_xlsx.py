import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\mhark\OneDrive\Documents\m88-dev-tracker\Untitled spreadsheet.xlsx')
ws = wb['Sheet1']

print(f"Max col: {ws.max_column}, Max row: {ws.max_row}")
print()

for col in range(1, ws.max_column + 1):
    h = ws.cell(row=1, column=col).value
    d = ws.cell(row=4, column=col).value
    fill = ws.cell(row=1, column=col).fill
    bg = None
    if fill and fill.fgColor:
        bg = fill.fgColor.rgb
    print(f"Col {col:>3} | BG={bg} | HEADER: {repr(h):<55} | DATA: {repr(d)}")
